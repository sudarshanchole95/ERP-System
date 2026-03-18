from __future__ import annotations
import sys
import json
import re
import hashlib
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from core.firestore_writer import FirestoreWriter
from core.validator import validate_documents

def fix_value(value):
    if value is None: return None
    
    # --- FIX: Convert Datetime to String (matches CSV format) ---
    if isinstance(value, datetime):
        # Converts to '2024-01-10T08:00:00' format
        return value.isoformat()
    
    if isinstance(value, (int, float)): 
        if isinstance(value, float) and value.is_integer(): return int(value)
        return value
    
    if isinstance(value, str):
        v = value.strip()
        if (v.startswith("{") and v.endswith("}")) or (v.startswith("[") and v.endswith("]")):
            try: return json.loads(v)
            except: pass
    return value

def generate_deterministic_id(row: dict) -> str:
    row_str = json.dumps(row, sort_keys=True, default=str)
    return hashlib.md5(row_str.encode("utf-8")).hexdigest()

def transform_row(row: dict) -> dict:
    row = {k: fix_value(v) for k, v in row.items()}

    address = {}
    for key in list(row.keys()):
        if key.startswith("address."):
            sub = key.split(".", 1)[1]
            address[sub] = row[key]
            del row[key]
    if address: row["address"] = address

    # Normalize empty
    for k, v in list(row.items()):
        if v == "" or v is None: row[k] = None

    # --- ID LOGIC ---
    if "id" in row and row["id"] is not None:
        row["id"] = str(row["id"])
    else:
        alias_id = next((row[k] for k in ["sku", "skuId", "code", "email", "storeId"] if k in row and row[k]), None)
        if alias_id:
            row["id"] = str(alias_id)
        else:
            row["id"] = generate_deterministic_id(row)

    return row

def excel_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    if not wb.sheetnames: return []
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
    except StopIteration: return []

    for row in rows:
        doc = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        yield transform_row(doc)

def import_excel_file(path, collection_name):
    docs = list(excel_rows(path))
    valid, errors = validate_documents(docs, collection_name)

    writer = FirestoreWriter()
    written, skipped = writer.write_documents(
        collection_name,
        valid,
        id_field="id",
        check_existing_and_skip_if_same=True
    )

    print(f"   📊 {Path(path).name:<30} -> {collection_name:<25} [Written: {written} | Skipped: {skipped}]")

if __name__ == "__main__":
    print("\n🚀 RUNNING STANDALONE EXCEL IMPORT...")
    excel_dir = BASE_DIR.parent / "sample-data" / "excel"
    
    if not excel_dir.exists():
        print(f"❌ Error: Directory not found at {excel_dir}")
        sys.exit(1)

    files = list(excel_dir.glob("*.xlsx"))
    if not files: print(f"⚠️ No Excel files found.")
    else:
        for f in files:
            coll = re.sub(r'[^a-zA-Z0-9]', '_', f.stem).strip('_')
            import_excel_file(str(f), coll)
    print("\n✅ Excel Import Complete.")