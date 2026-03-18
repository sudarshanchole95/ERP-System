import os
import re
import json
import csv
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook

# =====================================================================
#   CONFIGURATION: DYNAMIC OUTPUT PATH
# =====================================================================
# This saves the data relative to where this script is running.
# No more hardcoded "C:/Users/..." paths that break on typos.
BASE_DIR = Path(__file__).parent.parent.resolve()
FIXED_OUTPUT_DIR = BASE_DIR / "sample-data"

# =====================================================================
#   UTIL: EXPAND NESTED KEYS
# =====================================================================
def expand_nested(record):
    """
    Converts: {"address.street": "NY"} -> {"address": {"street": "NY"}}
    """
    result = {}
    for key, value in record.items():
        if "." not in key:
            result[key] = value
            continue
        parts = key.split(".")
        base = result
        for p in parts[:-1]:
            base = base.setdefault(p, {})
        base[parts[-1]] = value
    return result

# =====================================================================
#   1. UI: FILE SELECTION ONLY
# =====================================================================
def select_js_file():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title="Select the .js Data File",
        filetypes=[("JavaScript Files", "*.js"), ("All Files", "*.*")]
    )
    root.destroy()
    return path

def ask_export_format():
    print("\nSelect Export Format:")
    print("1. CSV only")
    print("2. Excel only")
    print("3. PDF only")
    print("4. ALL Formats (CSV + Excel + PDF)")
    
    choice = input("\nEnter choice (1-4): ").strip()
    return choice

# =====================================================================
#   2. PARSING LOGIC
# =====================================================================

def read_js_file(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def find_all_collections(js_text):
    matches = re.findall(r"\bconst\s+([a-zA-Z0-9_]+)\s*=\s*\[", js_text)
    return matches

def extract_js_block(js_text, name):
    pattern = re.compile(rf"\bconst\s+{re.escape(name)}\s*=\s*\[")
    match = pattern.search(js_text)
    if not match:
        return None

    start = match.end() - 1
    stack = ["["]
    i = start + 1

    while i < len(js_text) and stack:
        ch = js_text[i]
        if ch == "[":
            stack.append("[")
        elif ch == "]":
            stack.pop()
        elif ch in ("'", '"', '`'):
            quote = ch
            i += 1
            while i < len(js_text):
                curr = js_text[i]
                if curr == quote:
                    if js_text[i-1] != "\\": 
                        break
                i += 1
        i += 1

    return js_text[start:i]

def normalize_js(block):
    if block is None: return "[]"
    s = block
    s = re.sub(r"//.*?$", "", s, flags=re.MULTILINE)
    s = re.sub(r"/\*.*?\*/", "", s, flags=re.DOTALL)
    s = s.replace("`", '"')
    s = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)(\s*):', r'\1"\2"\3:', s)
    s = re.sub(r":\s*'([^']*)'", r': "\1"', s)
    s = re.sub(r",\s*([}\]])", r"\1", s)
    return s

def parse_collection(block):
    clean = normalize_js(block)
    try:
        data = json.loads(clean)
        return [expand_nested(r) for r in data]
    except json.JSONDecodeError as e:
        print(f"    ⚠️ Warning: JSON parsing error (attempting recovery): {e}")
        return []

# =====================================================================
#   3. EXPORT HELPERS
# =====================================================================
def flatten_value(v):
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    if v is None:
        return ""
    return str(v)

def export_csv(name, records, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{name}.csv"
    
    if not records: return
    
    cols = sorted({k for r in records for k in r})
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(cols)
        for r in records:
            writer.writerow([flatten_value(r.get(c, "")) for c in cols])
    print(f"    ✔ CSV: {path}")

def export_excel(name, records, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{name}.xlsx"
    
    if not records: return

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]
    cols = sorted({k for r in records for k in r})
    ws.append(cols)
    for r in records:
        ws.append([flatten_value(r.get(c, "")) for c in cols])
    wb.save(path)
    print(f"    ✔ Excel: {path}")

def export_pdf(name, records, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{name}.pdf"
    
    if not records: return

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, f"{name} Collection", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    for idx, rec in enumerate(records, start=1):
        json_block = json.dumps(rec, indent=2, ensure_ascii=False)
        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 8, f"Record {idx}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font("courier", size=9)
        pdf.multi_cell(0, 5, json_block, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(2)

    pdf.output(path)
    print(f"    ✔ PDF: {path}")

# =====================================================================
#   MAIN PIPELINE
# =====================================================================

def main():
    print("🚀 JS Data Extractor Started")
    
    # 1. Ask User for Input File
    print("\n👉 Step 1: Select your .js file...")
    js_path = select_js_file()
    if not js_path:
        print("❌ No file selected. Exiting.")
        return
    print(f"   Input File: {js_path}")

    # 2. Use Safe Relative Path
    print(f"\n👉 Step 2: Target Folder set to:")
    print(f"   {FIXED_OUTPUT_DIR}")
    
    # Ensure base directory exists
    FIXED_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 3. Select Format
    fmt_choice = ask_export_format()
    
    # 4. Processing
    js_text = read_js_file(js_path)
    collections = find_all_collections(js_text)
    
    if not collections:
        print("❌ No data arrays found.")
        return

    print(f"\n🔍 Found {len(collections)} collections: {', '.join(collections)}\n")

    for name in collections:
        print(f"▶ Processing '{name}'...")
        block = extract_js_block(js_text, name)
        parsed_records = parse_collection(block)
        
        if not parsed_records:
            print(f"    ⚠️ No records found for '{name}'")
            continue

        if fmt_choice == '1' or fmt_choice == '4':
            export_csv(name, parsed_records, FIXED_OUTPUT_DIR / "csv")
        
        if fmt_choice == '2' or fmt_choice == '4':
            export_excel(name, parsed_records, FIXED_OUTPUT_DIR / "excel")

        if fmt_choice == '3' or fmt_choice == '4':
            export_pdf(name, parsed_records, FIXED_OUTPUT_DIR / "pdf")
            
        print("-" * 30)

    print(f"\n✅ All tasks completed!")
    print(f"📂 Files saved in: {FIXED_OUTPUT_DIR}")

if __name__ == "__main__":
    main()